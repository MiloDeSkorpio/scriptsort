##importar librearias necesarias para el funcionamiento del script
import os
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
## Nombre del mes con texto, se ocupara para leer la carpeta del mes y asignar el nombre a los archivos generados
mes_nombre = "Febrero"

## Modificar el contenido de m = "mes" * Para los meses que anteriores a octubre ocupar la sintaxis 09 = Septiembre 08 = Agosto
## Modificar el contenido de Y = "Año" 2023 / 2024 / 2025 
m = "02"
y = "2024"

##
semana = "6"
## Nombre de las extenciones de los archivos que ocupara el script para realizar 
a = "-Transacciones.csv"
ae = "-Transacciones-extension.csv"

## Ruta de la cual se extraeran todos los archivos y en la misma se guardaran los archivos
ruta_guardado = f"Transacciones/{y}/{m} {mes_nombre}"

## Este es el rango de dias en el que se trabajara, para el tema del ultimo dia siempre se le sumara 1
## Ejemplo primera quincena dia_fn = 16 el metodo range trabaja de esa forma
dia_in = 13
dia_fn = 14
rango = dia_fn - dia_in

## Listado de los archvios -Transacciones.csv
## Listado de los archivo a leer segun el rango especificado 
archivo_tr = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{a}") for d in range(dia_in, dia_fn)]

## Leer Archivos de Extencion para obtener la duracion de las transacciones
## Lista de nombres de archivo
archivo_ex = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{ae}") for d in range(dia_in, dia_fn)]

## Areglo que se llenara con los archivos -Transacciones-extension.csv
transacciones = []
## Bucle para insertar todos los archivos en el DataFrame transacciones
for transaccion in archivo_tr:
  df = pd.read_csv(transaccion, low_memory=False)
  transacciones.append(df)

## Concatenación de documentos extraídos del arreglo transacciones y creando un solo DataFrame con información de toda la quincena
## Arreglo que se llenara con los archivos -Transacciones.csv
extenciones = []
## Bucle para insertar todos los archivos en el DataFrame extenciones
for extencion in archivo_ex:
  df = pd.read_csv(extencion)
  extenciones.append(df)  

## Concatenación de documentos extraídos del arreglo extenciones y creando un solo DataFrame con información de todo el mes
df_extenciones = pd.concat(extenciones, ignore_index=True)

resumen = []

## Funcion para el analisis general
def resumen_transacciones(transacciones):
  """
  Función para generar un resumen de transacciones por tipo y método.
  Args:
    transacciones: Lista de DataFrames de transacciones.
  Returns:
    Lista de diccionarios con el resumen de transacciones.
  """
  for df in transacciones:
    ## Filtrar las transacciones de tipo 0
    df['TIPO_TRANSACCION'] = df['TIPO_TRANSACCION'].astype('str')
    df_filtro = df[df['TIPO_TRANSACCION'] == '0'].copy()

    ## Sacaremos el total de transacciones con el metodo count 
    tr_totales = df_filtro['TIPO_TRANSACCION'].count()

    ## Convertir la columna FECHA_HORA_TRANSACCION a datetime
    df_filtro['FECHA_HORA_TRANSACCION'] = pd.to_datetime(df_filtro['FECHA_HORA_TRANSACCION'])
    df_filtro['FECHA_HORA_TRANSACCION'] = df_filtro['FECHA_HORA_TRANSACCION'].dt.strftime('%Y-%m-%d')

    ## Separar las transacciones en base al método
    df_fisico = df_filtro[df_filtro['LOCATION_ID'] == '201A00']
    df_digital = df_filtro[df_filtro['LOCATION_ID'] == '101800']
    df_appcdmx = df_filtro[df_filtro['LOCATION_ID'] == '101801']

    ## Calcular el monto total por transacción física y agregar al DataFrame correspondiente
    monto_fisico = df_fisico['MONTO_TRANSACCION'].sum()

    ## Calcular el monto total por transacción digital y agregar al DataFrame correspondiente
    monto_digital = df_digital['MONTO_TRANSACCION'].sum()
    ## 
    monto_appcdmx = df_appcdmx['MONTO_TRANSACCION'].sum()

    ## Obtener los valores únicos de la fecha de transacción
    fechas_unicas = df_filtro['FECHA_HORA_TRANSACCION'].unique()

    ## Agregar los resultados a la lista de resumen 
    resumen.append({
        'FECHA': ', '.join(fechas_unicas),
        'TR Digitales': df_digital.shape[0],
        'TR Fisicas': df_fisico.shape[0],
        'TR AppCDMX': df_appcdmx.shape[0],
        'TR Totales': tr_totales,
        'Montos Digitales': monto_digital / 100,
        'Montos Fisicos': monto_fisico / 100,
        'Montos AppCDMX': monto_appcdmx / 100,
        'Monto Total': (monto_digital + monto_fisico + monto_appcdmx)/100,
    })
  return resumen
## Definir funcion para crear graficos:

def crear_grafico(rango,resultados,title,nombre_grafico,width):
    ## Grafico
    print("Creando Grafico")
    ## Definir el estilo del grafico
    plt.style.use('seaborn-v0_8-paper')
    ## Definir el ancho de cada una de las barras
    width = width  
    ## Definimos el objeto grafico (ancho,alto)
    fig,ax = plt.subplots(figsize=(18,8))
    ## Valores base para las barras apiladas
    bottom = np.zeros(rango)
    ## Colores para las barras
    colors = ['#385723','#a50021'] 
    # Define variables
    fechas = resultados['FECHA']
    tr_rrd = resultados['TR Digitales'] + resultados['TR AppCDMX']
    tr_rrf = resultados['TR Fisicas']
    mto_tt = resultados['Monto Total']
    # Asigna nombres a las Series
    tr_rrd.name = 'TR Digitales'
    tr_rrf.name = 'TR Fisicas'
    mto_tt.name = 'Monto Total'
    # Crea un DataFrame
    tr_counts = pd.DataFrame({
        'TR Digitales': np.array(tr_rrd),
        'TR Fisicas': np.array(tr_rrf),
    })
    ## Creamos la iteracion de los datos
    for i, (tr, tr_count) in enumerate(tr_counts.items()):
        ## Creamos un grafico de barras
        p = ax.bar(fechas, tr_count, width, label=tr, bottom=bottom, color=colors[i])
        ## Generamos el apilamiento de barras
        bottom += tr_count
        ## Asignamos las etiwuetas a las barras
        ax.bar_label(p, label_type='center', color='#fff',fontsize=10,fontweight=600,**{'fmt': '{:,.0f}'})
    ## Creamos un segundo eje
    ax2 = ax.twinx()
    ## Graficamos una linea con el nuevo eje
    ax2.plot(fechas, mto_tt, label="Montos", color="#fe9c55", marker='o',linestyle="solid")
    ## Añadimos las etiquetas a la linea
    for i, (fecha, monto) in enumerate(zip(fechas, mto_tt)):
        if monto >= min(mto_tt) :  
            ax2.annotate(f"${monto:,.0f}", (fecha, monto), xytext=(2,5), textcoords='offset points', fontsize=10,fontweight=600)

    ## Posicionamiento de la etiqueta del eje secundario
    ax2.yaxis.set_label_position("right")
    ## Se asignan las etiqetas para los ejes
    ax2.set_ylabel("Valor Monetario",fontsize=8,fontweight=600)
    ax.set_ylabel("N° de Transacciones",fontsize=8,fontweight=600)
    ax.set_xlabel(f'Semana {semana}',fontsize=8,fontweight=600)
    ## Creamos una legenda fuera del grafico
    fig.legend(loc='outside upper left')
    # Ajusta el formato de los valores en el eje Y
    ax.yaxis.set_major_formatter(lambda x, pos: f'{x:,.0f}')
    ax2.yaxis.set_major_formatter(lambda x, pos: f'{x:,.0f}')
    # Ajusta el título del gráfico
    ax.set_title(title,fontsize=12,fontweight=600)
    # Guarda el gráfico en alta resolución
    ruta_grafico = os.path.join(ruta_guardado, nombre_grafico)
    plt.savefig(ruta_grafico,format='png',dpi=900,bbox_inches='tight')
    print("Proceso realizado con Exito!!")
## Inicia el condicional para los dias sueltos
if rango < 13 :
  ## Bucle para el analisis de todas las transacciones
  resumen_transacciones(transacciones)
  ## Convertir el arreglo resumen en DataFrame
  resultados = pd.DataFrame(resumen)
  ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
  archivo_sem = f"RRE_{mes_nombre}_{dia_in}-{dia_fn-1}.csv"
  ruta_res_sem = os.path.join(ruta_guardado, archivo_sem)
  resultados.to_csv(ruta_res_sem, index=False)
  print("Proceso realizado con Exito!!")
  if rango == 7 :
    nombre_grafico = f'RR_Grafico_Semana_{semana}.png'
    title = f'Montos recaudados por tipo de red semana {semana}'
    width = 0.6
    print(f"Trabajando con la semana {semana}")
    crear_grafico(rango,resultados,title,nombre_grafico,width)
    
## Inicia el condicional para las Quincenas        
elif rango >= 13 and rango <= 16:
  print("Realizando analisis quincenal.")
  ## Nombres de las quincenas
  first = "1ra"
  second = "2da"

  ## Concatenar el arreglo
  df_transacciones = pd.concat(transacciones, ignore_index=True)
  
  ## Arreglo para almacenar el resumen de las tarjetas
  tarjetas = []

  # Filtrar las transacciones de tipo 0
  df_filtered = df_transacciones[df_transacciones['TIPO_TRANSACCION'] == '0'].copy()

  # Obtener el monto total
  monto = df_filtered['MONTO_TRANSACCION'] / 100
  
  # Obtener las tarjetas únicas
  tarjetas_unicas = df_filtered['NUMERO_SERIE_HEX'].unique()

  # Agregar los resultados a la lista de tarjetas
  tarjetas.append({
    '# Tarjetas': len(tarjetas_unicas),
    '# Transacciones': len(df_filtered),
    '$ Promedio': monto.mean(),
  })

  if dia_in == 1:
    print(f"Analizando la {first} qna de {mes_nombre}")
    ## Creando RRE
    print(f"Generando RRE {first} qna {mes_nombre}")
    resumen_transacciones(transacciones)
    ## Convertir el arreglo resumen en DataFrame
    resultados = pd.DataFrame(resumen)
    ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
    archivo_mens = f"RRE_{first}_qna_{mes_nombre}.csv"
    ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
    resultados.to_csv(ruta_res_mens, index=False)
    ## Archivo completo
    print("Generando datos en crudo")
    archivo_full = f"Full_{first}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ##   Ectenciones
    print("Generando datos extenciones")
    archivo_full = f"Full_ext_{first}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ## Resumen de Tarjetas
    print("Generando resumen de las tarjetas")
    res_tarjetas = pd.DataFrame(tarjetas)
    archivo_tar = f"Tarjetas_{first}_qna_{mes_nombre}.csv"
    ruta_resultados = os.path.join(ruta_guardado, archivo_tar)
    res_tarjetas.to_csv(ruta_resultados, index=False)
    
    ## Grafico
    title = f'Transacciones por tipo de red por día del {dia_in} al {dia_fn -1} de enero'
    nombre_grafico = f'RR_Grafico_{first}_qna.png'
    width = 0.8
    crear_grafico(rango,resultados,title,nombre_grafico,width)
  ## Condicion para la segunda Quincena
  elif dia_in == 16:
    print(f"Analizando la {second} qna de {mes_nombre}")
    ## Creando RRE
    print(f"Generando RRE {second} qna {mes_nombre}")
    resumen_transacciones(transacciones)
    ## Convertir el arreglo resumen en DataFrame
    resultados = pd.DataFrame(resumen)
    ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
    archivo_mens = f"RRE_{second}_qna_{mes_nombre}.csv"
    ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
    resultados.to_csv(ruta_res_mens, index=False)
    ## Archivo completo 
    print("Generando datos en crudo")
    archivo_full = f"Full_{second}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ## Extenciones    
    print("Generando datos extenciones")
    archivo_full = f"Full_ext_{second}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ## Resumen de tarjetas
    print("Generando resumen de las tarjetas")
    res_tarjetas = pd.DataFrame(tarjetas)
    archivo_tar = f"Tarjetas_{second}_qna_{mes_nombre}.csv"
    ruta_resultados = os.path.join(ruta_guardado, archivo_tar)
    res_tarjetas.to_csv(ruta_resultados, index=False)
    ## Grafico
    nombre_grafico = f'RR_Grafico_{second}_qna.png'
    title = f'Transacciones por tipo de red por día del {dia_in} al {dia_fn -1} de enero'
    width = 0.8
    crear_grafico(rango,resultados,title,nombre_grafico, width)
## Inicia el Condicional para los meses    
elif rango > 16:
  ## Se guarda la concatenacion de todo el mes para conciliacion con SEMOVI
  df_transacciones = pd.concat(transacciones, ignore_index=True)
  print("Generando datos en crudo")
  archivo_full = f"Full_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_transacciones.to_csv(ruta_full, index=False)
  ## Extenciones    
  print("Generando datos extenciones")
  archivo_full = f"Full_ext_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_transacciones.to_csv(ruta_full, index=False)
  resumen_transacciones(transacciones)
  ## llamamos a la funcion resumen_transacciones 
  ## Convertir el arreglo resumen en DataFrame
  resultados = pd.DataFrame(resumen)
  ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
  archivo_mens = f"RRE_{mes_nombre}.csv"
  ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
  resultados.to_csv(ruta_res_mens, index=False)
  ## Aqui inicia el proceso para el analisis del documento de mercado pago
  ########################################################################
  ## Totales de transacciones fisicas para poder realizar operaciones
  tr_fisicas = resultados['TR Fisicas'].sum()
  tr_digitales = resultados['TR Digitales'].sum()
  tr_appcdmx = resultados['TR AppCDMX'].sum()
  tr_total = resultados['TR Totales'].sum()

  ## Obtener datos porcentuales de transacciones fisicas, digitales y el total
  pr_dig = tr_digitales / tr_total  
  pr_fis = tr_fisicas / tr_total  
  pr_app = tr_appcdmx / tr_total  
  pr_total = pr_dig + pr_fis + pr_app

  ## Tabla de valores porcentuales para la comision de mercado pago
  ## Recargas Digitales
  ## Porcentaje mensual de comision digital
  pr_dg1 = 2.2
  pr_dg2 = 2.1
  pr_dg3 = 2
  pr_dg4 = 1.9

  ## Recargas Fisicas (Negocios)
  ## Porcentaje mensual de comision fisica
  pr_fs1 = 2
  pr_fs2 = 1.9
  pr_fs3 = 1.8
  pr_fs4 = 1.7
  
  ## Porcentaje mensual de comision app cdmx
  pr_ap1 = 2
  pr_ap2 = 1.9
  pr_ap3 = 1.8
  pr_ap4 = 1.7

  ## Limites de Rango, esta es la tabla que se ecnuentra en el presente contrato de Mercado Pago
  ## Los mismos rangos se utilizan para recargas fisicas y digitales 
  r1 = 15000000
  r2 = 30000000
  r3 = 45000000
  r4 = 60000000

  ## Obtener totales en $ para fisicas, digitales y el total de la suma de ambas 
  tt_fisico = resultados['Montos Fisicos'].sum()
  tt_digital = resultados['Montos Digitales'].sum()
  tt_appcdmx = resultados['Montos AppCDMX'].sum()
  mt_total = tt_fisico + tt_digital + tt_appcdmx

  ## Se realizan las condicionales para ajustar el porcentaje automaticamente segun el total segun sea el caso fisica o digital  
  ## Condicional de porcentaje Digital
  if tt_digital <= r1:
      pr_com_dig = pr_dg1
  elif tt_digital <= r2:
      pr_com_dig = pr_dg2
  elif tt_digital <= r3:
      pr_com_dig = pr_dg3
  elif tt_digital <= r4:
      pr_com_dig = pr_dg4
      
  ## Condicional de porcentaje Fisico
  if tt_fisico <= r1:
      pr_com_fis = pr_fs1
  elif tt_fisico <= r2:
      pr_com_fis = pr_fs2
  elif tt_fisico <= r3:
      pr_com_fis = pr_fs3
  elif tt_fisico <= r4:
      pr_com_fis = pr_fs4
      
  ## Condicional de appcdmx
  if tt_appcdmx <= r1:
      pr_com_app = pr_ap1
  elif tt_appcdmx <= r2:
      pr_com_app = pr_ap2
  elif tt_appcdmx <= r3:
      pr_com_app = pr_ap3
  elif tt_appcdmx <= r4:
      pr_com_app = pr_ap4

  ## Comisiones Fisicas y Digitales
  com_fisico = (pr_com_fis/100)*tt_fisico
  com_digital = (pr_com_dig/100)*tt_digital
  com_appcdmx = (pr_com_app/100)*tt_appcdmx
  com_total = com_fisico + com_digital + com_appcdmx
  prm_digital = tt_digital / mt_total 
  prm_fisico = tt_fisico / mt_total 
  prm_appcdmx = tt_appcdmx / mt_total 
  prm_total = prm_digital + prm_fisico + prm_appcdmx
  
  #Resultados Transacciones 
  wb = Workbook()

  ## Si existe una hoja llamada Sheet, se elimina para evitar crear una hoja vacia
  if 'Sheet' in wb.sheetnames:
      wb.remove(wb['Sheet'])
      
  ## Declaramos las hojas en las cuales se van a guardar todas las tablas y guardar la informacion
  hoja = wb.create_sheet(title="Reporte Mensual MP")

  ## Agregar la informacion del total de transacciones y el porcentaje total que ocupan tanto fisicas y digitales
  ## Se ocupara la hoja 1 para guardar las tablas de montos y transacciones totales con sus porcentajes
  hoja['A1'] = "Tabla de Transacciones"
  hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
  hoja.append(['Cantidad de Recargas',tr_digitales,tr_fisicas,tr_appcdmx,tr_total])
  hoja.append(['Proporcion',pr_dig,pr_fis,pr_app,pr_total])

  ## Crear un campo vacio para darle un espacio entre la primer y segunda tabla
  hoja.append([])

  ## Tabla con la informacion de el total de montos con sus porcentajes para fisicas y digitales en base al rango que ocupen
  hoja['A6'] = "Tabla de Montos y Proporciones"
  hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
  hoja.append(['Cantidad de Recargas',tt_digital,tt_fisico,tt_appcdmx,mt_total])
  hoja.append(['Proporcion',prm_digital,prm_fisico,prm_appcdmx,prm_total])
  hoja.append(['Comision para MP',com_digital,com_fisico,com_appcdmx,com_total])
  ## Se guarda el archivo de Mercado pago
  archivo_mp = f"Reporte_MP_{mes_nombre}.xlsx"
  ruta_mp = os.path.join(ruta_guardado,archivo_mp)
  wb.save(ruta_mp)
  
  ## Aqui inicia el proceso para el analisis de las penalizaciones de transacciones mayores a 7 segundos
  ########################################################################
  ## Analisis para las penalizaciones de las transacciones mayores a 7 seg
  mayor_seven = df_extenciones['DURATION'] > 7
  
  ## Se realiza el conteo total de todas las transacciones que son mayores a los 7 segundos
  ntr_may_seven = df_extenciones.loc[mayor_seven,['DURATION']].count()
  
  ## Convierte la serie a un tipo de datos numérico
  list_mayor_seven = df_extenciones.loc[mayor_seven,['ID_TRANSACCION_ORGANISMO', 'DURATION','END_DATE']]
  
  ## Convierte la serie a un tipo de datos numérico
  df_merge_succ = pd.merge(list_mayor_seven, df_transacciones, on='ID_TRANSACCION_ORGANISMO', how='inner')
  df_merge_fil = df_merge_succ[['ID_TRANSACCION_ORGANISMO', 'LOCATION_ID', 'MONTO_TRANSACCION', 'END_DATE', 'DURATION']]
  df_merge_fil['MONTO_TRANSACCION'] = df_merge_fil['MONTO_TRANSACCION'].apply(lambda x: x / 100)

  lista_tr_7s = f"RRE - Penalizaciones {mes_nombre}.xlsx"
  ruta_lista = os.path.join(ruta_guardado,lista_tr_7s)
  with pd.ExcelWriter(ruta_lista) as writer:
      df_merge_fil.to_excel(writer, index=False ,sheet_name=f'Transacciones penalizables {mes_nombre}')

  print("Proceso realizado con Exito!!")
  