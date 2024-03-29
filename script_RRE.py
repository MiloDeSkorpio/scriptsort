##importar librearias necesarias para el funcionamiento del script
import os
import pandas as pd
from openpyxl import Workbook

## Nombre del mes con texto, se ocupara para leer la carpeta del mes y asignar el nombre a los archivos generados
mes_nombre = "Enero"

## Modificar el contenido de m = "mes" * Para los meses que anteriores a octubre ocupar la sintaxis 09 = Septiembre 08 = Agosto
## Modificar el contenido de Y = "Año" 2023 / 2024 / 2025 
m = "01"
y = "2024"

## Nombre de las extenciones de los archivos que ocupara el script para realizar 
a = "-Transacciones.csv"
ae = "-Transacciones-extension.csv"
at = "-Test-Transacciones.csv"

## Ruta de la cual se extraeran todos los archivos y en la misma se guardaran los archivos
ruta_guardado = f"Transacciones/{y}/{m} {mes_nombre}"

## Quincena a trabajar
first = "1ra"
second = "2da"

## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
mes_filtrado = f"{mes_nombre}_filtrado.csv"
mes_completo = f"{mes_nombre}_completo.csv"

## Archivos quincenales agregar # al inicio cuando se desea trabajar por mes el script
#archivo_mp = f"Reporte_MP_{second}_qna_{mes_nombre}.xlsx"
#quincena = f"{second}_qna_{mes_nombre}"

## Archivos mensuales eliminar # cuando se trabaje el script por mes
archivo_mp = f"Reporte_MP_{mes_nombre}.xlsx"## Desactivar cuando se requiera el Reporte_MP_mensual
quincena = f"{mes_nombre}" ## Desactivar cuando se requiera el Resumen_RRE_mensual

## Este es el rango de dias en el que se trabajara, para el tema del ultimo dia siempre se le sumara 1
## Ejemplo primera quincena dia_fn = 16 el metodo range trabaja de esa forma
dia_in =  1
dia_fn = 16

## Listado de los archvios -Transacciones.csv
## Listado de los archivo a leer segun el rango especificado 
archivo_tr = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{a}") for d in range(dia_in, dia_fn)]

## Leer Archivos de Extencion para obtener la duracion de las transacciones
## Lista de nombres de archivo
archivo_ex = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{ae}") for d in range(dia_in, dia_fn)]

## Areglo que se llenara con los archivos -Transacciones-extension.csv
extenciones = []

## Arreglo que se llenara con los archivos -Transacciones.csv
transacciones = []

## Arreglo que recopilara los datos Resumidos del filtro y se ocupara para realizar el RRE
resumen = []

## Bucle para insertar todos los archivos en el DataFrame transacciones

for transaccion in archivo_tr:
    df = pd.read_csv(transaccion)
    transacciones.append(df)

## Bucle para insertar todos los archivos en el DataFrame extenciones
for extencion in archivo_ex:
    df = pd.read_csv(extencion)
    extenciones.append(df)   
    
## Concatenación de documentos extraídos del arreglo transacciones y creando un solo DataFrame con información de toda la quincena
df_completo = pd.concat(transacciones, ignore_index=True)

## Concatenación de documentos extraídos del arreglo extenciones y creando un solo DataFrame con información de todo el mes
mes = pd.concat(extenciones, ignore_index=True)

archivo_full = f"1ra_qna_ene_ext.csv"
ruta_full = os.path.join(ruta_guardado, archivo_full)
mes.to_csv(ruta_full, index=False)

## Serie_Hex de tarjetas unicas utilizando un filtro para solo obtener transacciones validas
df_completo['TIPO_TRANSACCION'] = df_completo['TIPO_TRANSACCION'].astype('str')
df_trfil = df_completo[df_completo['TIPO_TRANSACCION'] == '0' ].copy()
df_trfil['TIPO_TRANSACCION'] = pd.to_numeric(df_trfil['TIPO_TRANSACCION'])

tarjetas_unicas = pd.DataFrame({'NUMERO_SERIE_HEX': df_trfil['NUMERO_SERIE_HEX'].unique()})

result = pd.merge(df_trfil, pd.DataFrame(tarjetas_unicas, columns=['NUMERO_SERIE_HEX']), on='NUMERO_SERIE_HEX', how='inner')

for column in ['LINEA','ESTACION','AUTOBUS','RUTA','EQUIPO','CONTADOR_VALIDACIONES','PURCHASE_LOG','COUNTER_VALUE','COUNTER_AMOUNT']:
    df_completo = df_completo.drop(column, axis=1)

for datos in ['DEVICE_ID','LATITUDE','LONGITUDE']:
    mes = mes.drop(datos,axis=1)    

empty_field = df_completo.isnull().sum()
mes_field = mes.isnull().sum()


num_trx = df_trfil['TIPO_TRANSACCION'].size 
## Crear DataFrame Tarjetas con todas las tarjetas unicas y poder compararla con las transacciones y asignarle el monto

## MONTOS
## Elimina duplicados en df_trfil basado en la columna 'NUMERO_SERIE_HEX'
df_trfil = df_trfil.drop_duplicates(subset='NUMERO_SERIE_HEX')

## Guardar el DataFrame de filtrado en un archivo CSV este contiene la concatenacion de los archivos
## con  extencion -Transacciones.csv pero con un filtro de que solo aparezan las transacciones con
## ['TIPO_TRANSACCION'] == '0'
ruta_filtrado = os.path.join(ruta_guardado, mes_filtrado)
df_trfil.to_csv(ruta_filtrado, index=False)

## Luego, realiza la combinación con tarjetas_unicas
recargos = result['MONTO_TRANSACCION']

## Utilizamos la variable recargas para dividirla entre 100 y concertir de centavos a pesos
montos = recargos/100
mts = df_trfil['MONTO_TRANSACCION']
mtsfp = mts/100

## A la variable de montos le sacaremos el promedio para poder pasarla al documento
mean_montos = montos.mean()
mean_mtsfp = mtsfp.mean()

## Bucle para el analisis de todas las transacciones '
for df in transacciones:
    ## Filtrar las transacciones de tipo 0
    df['TIPO_TRANSACCION'] = df['TIPO_TRANSACCION'].astype('str')
    df_filtro = df[df['TIPO_TRANSACCION'] == '0'].copy()
    df_filtro['TIPO_TRANSACCION'] = pd.to_numeric(df_filtro['TIPO_TRANSACCION'])
    ## Sacaremos el total de transacciones con el metodo count 
    tr_totales = df_filtro['TIPO_TRANSACCION'].count()
    
    ## Convertir la columna FECHA_HORA_TRANSACCION a datetime
    df_filtro['FECHA_HORA_TRANSACCION'] = pd.to_datetime(df_filtro['FECHA_HORA_TRANSACCION'])
    df_filtro['FECHA_HORA_TRANSACCION'] = df_filtro['FECHA_HORA_TRANSACCION'].dt.strftime('%Y-%m-%d')

    ## Separar las transacciones en base al método
    df_fisico = df_filtro[df_filtro['LOCATION_ID'] == '201A00']
    df_digital = df_filtro[df_filtro['LOCATION_ID'] == '101800']
    df_appcdmx = df_filtro[df_filtro['LOCATION_ID'] == '101801']
    
    ## Calcular el monto total por transacción física y agregar al DataFrame correspondiente
    monto_fisico = df_fisico['MONTO_TRANSACCION'].sum()
    
    ## Calcular el monto total por transacción digital y agregar al DataFrame correspondiente
    monto_digital = df_digital['MONTO_TRANSACCION'].sum()
    ## 
    monto_appcdmx = df_appcdmx['MONTO_TRANSACCION'].sum()
    
    ## Obtener los valores únicos de la fecha de transacción
    fechas_unicas = df_filtro['FECHA_HORA_TRANSACCION'].unique()

    ## Agregar los resultados a la lista de resumen 
    resumen.append({
        'FECHA': ', '.join(fechas_unicas),
        'TR Digitales': df_digital.shape[0],
        'TR Fisicas': df_fisico.shape[0],
        'TR AppCDMX': df_appcdmx.shape[0],
        'TR Totales': tr_totales,
        'Montos Digitales': monto_digital / 100,
        'Montos Fisicos': monto_fisico / 100,
        'Montos AppCDMX': monto_appcdmx / 100,
        'Monto Total': (monto_digital + monto_fisico + monto_appcdmx)/100,
    })

## Convertir el arreglo resumen en DataFrame
resultados = pd.DataFrame(resumen)

## Totales de transacciones fisicas para poder realizar operaciones
tr_fisicas = resultados['TR Fisicas'].sum()
tr_digitales = resultados['TR Digitales'].sum()
tr_appcdmx = resultados['TR AppCDMX'].sum()
tr_total = resultados['TR Totales'].sum()

## Obtener datos porcentuales de transacciones fisicas, digitales y el total
pr_dig = tr_digitales / tr_total * 100 
pr_fis = tr_fisicas / tr_total * 100 
pr_app = tr_appcdmx / tr_total * 100 
pr_total = pr_dig + pr_fis + pr_app

## Tabla de valores porcentuales para la comision de mercado pago
## Recargas Digitales
## Porcentaje mensual de comision digital
pr_dg1 = 2.2
pr_dg2 = 2.1
pr_dg3 = 2
pr_dg4 = 1.9

## Recargas Fisicas (Negocios)
## Porcentaje mensual de comision fisica
pr_fs1 = 2
pr_fs2 = 1.9
pr_fs3 = 1.8
pr_fs4 = 1.7
## Porcentaje mensual de comision app cdmx
pr_ap1 = 2
pr_ap2 = 1.9
pr_ap3 = 1.8
pr_ap4 = 1.7

## Limites de Rango, esta es la tabla que se ecnuentra en el presente contrato de Mercado Pago
## Los mismos rangos se utilizan para recargas fisicas y digitales 
r1 = 15000000
r2 = 30000000
r3 = 45000000
r4 = 60000000

## Obtener totales en $ para fisicas, digitales y el total de la suma de ambas 
tt_fisico = resultados['Montos Fisicos'].sum()
tt_digital = resultados['Montos Digitales'].sum()
tt_appcdmx = resultados['Montos AppCDMX'].sum()
mt_total = tt_fisico + tt_digital + tt_appcdmx

## Se realizan las condicionales para ajustar el porcentaje automaticamente segun el total segun sea el caso fisica o digital  
## Condicional de porcentaje Digital
if tt_digital <= r1:
    pr_com_dig = pr_dg1
elif tt_digital <= r2:
    pr_com_dig = pr_dg2
elif tt_digital <= r3:
    pr_com_dig = pr_dg3
elif tt_digital <= r4:
    pr_com_dig = pr_dg4
    
## Condicional de porcentaje Fisico
if tt_fisico <= r1:
    pr_com_fis = pr_fs1
elif tt_fisico <= r2:
    pr_com_fis = pr_fs2
elif tt_fisico <= r3:
    pr_com_fis = pr_fs3
elif tt_fisico <= r4:
    pr_com_fis = pr_fs4
    
## Condicional de appcdmx
if tt_appcdmx <= r1:
    pr_com_app = pr_ap1
elif tt_appcdmx <= r2:
    pr_com_app = pr_ap2
elif tt_appcdmx <= r3:
    pr_com_app = pr_ap3
elif tt_appcdmx <= r4:
    pr_com_app = pr_ap4

## Comisiones Fisicas y Digitales
com_fisico = (pr_com_fis/100)*tt_fisico
com_digital = (pr_com_dig/100)*tt_digital
com_appcdmx = (pr_com_app/100)*tt_appcdmx
com_total = com_fisico + com_digital + com_appcdmx
prm_digital = tt_digital / mt_total *100
prm_fisico = tt_fisico / mt_total *100
prm_appcdmx = tt_appcdmx / mt_total *100
prm_total = prm_digital + prm_fisico + prm_appcdmx

## Este es el filtro del cual se obtienen las transacciones que tienen una duracion mayor a 7 segundos 
mayor_seven = mes['DURATION']>7

## resumen

resumen = mes['DURATION'].value_counts()
resumen.name = '# de transacciones'
resumen.index.name ='Tiempo'
#resumen = resumen.rename(column={'DURATION': 'Tiempo'})
## Se realiza el conteo total de todas las transacciones que son amyores a los 7 segundos
ntr_may_seven = mes.loc[mayor_seven,['DURATION']].count()

## Convierte la serie a un tipo de datos numérico
list_mayor_seven = mes.loc[mayor_seven,['ID_TRANSACCION_ORGANISMO', 'DURATION','END_DATE']]
## Convierte la serie a un tipo de datos numérico
df_merge_succ = pd.merge(list_mayor_seven, df_completo, on='ID_TRANSACCION_ORGANISMO', how='inner')
df_merge_fil = df_merge_succ[['ID_TRANSACCION_ORGANISMO', 'LOCATION_ID', 'MONTO_TRANSACCION', 'END_DATE', 'DURATION']]
df_merge_fil['MONTO_TRANSACCION'] = df_merge_fil['MONTO_TRANSACCION'].apply(lambda x: x / 100)

lista_tr_7s = f"RRE - Penalizaciones {mes_nombre}.xlsx"
ruta_lista = os.path.join(ruta_guardado,lista_tr_7s)

#res_mayor_sev = resultados.drop(['TR Fisicas','TR Digitales', 'TR Totales'], axis=1)
#res_mayor_sev.columns = ['FECHA', 'RRFisica', 'RRDigital', 'RRE']

with pd.ExcelWriter(ruta_lista) as writer:
    #res_mayor_sev.to_excel(writer, index=False ,sheet_name=f'{mes_nombre}')
    df_merge_fil.to_excel(writer, index=False ,sheet_name=f'Transacciones penalizables {mes_nombre}')

#Resultados Transacciones 
wb = Workbook()

## Si existe una hoja llamada Sheet, se elimina para evitar crear una hoja vacia
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
    
## Declaramos las hojas en las cuales se van a guardar todas las tablas y guardar la informacion
hoja = wb.create_sheet(title="Reporte Mensual MP")
hoja2 = wb.create_sheet(title="TR Mayores a 7s")
hoja3 = wb.create_sheet(title="Promedio y Tarjetas")

## Agregar la informacion del total de transacciones y el porcentaje total que ocupan tanto fisicas y digitales
## Se ocupara la hoja 1 para guardar las tablas de montos y transacciones totales con sus porcentajes
hoja['A1'] = "Tabla de Transacciones"
hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
hoja.append(['Cantidad de Recargas',tr_digitales,tr_fisicas,tr_appcdmx,tr_total])
hoja.append(['Proporcion',pr_dig,pr_fis,pr_app,pr_total])

## Crear un campo vacio para darle un espacio entre la primer y segunda tabla
hoja.append([])

## Tabla con la informacion de el total de montos con sus porcentajes para fisicas y digitales en base al rango que ocupen
hoja['A6'] = "Tabla de Montos y Proporciones"
hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
hoja.append(['Cantidad de Recargas',tt_digital,tt_fisico,tt_appcdmx,mt_total])
hoja.append(['Proporcion',prm_digital,prm_fisico,prm_appcdmx,prm_total])
hoja.append(['Comision para MP',com_digital,com_fisico,com_appcdmx,com_total])

## En la hoja 2 se guardaran el N° total de transacciones que tienen una duracion mayor a 7 segundos
hoja2['A1'] = "Total de Transacciones Mayores a 7 Segundos"
hoja2.append(['Total',ntr_may_seven.to_string()])

## Crear un campo vacio para darle un espacio entre la primer y segunda tabla
hoja.append([])

## Crear la tabla de todas las transacciones mayores a 7 segundos con el ID_TRANSACCION_ORGANISMO
hoja2['A4'] = "Lista de Transacciones Mayores a 7 Segundos"

## Crear una lista de encabezados de columnas (opcional)
column_headers = ['ID_TRANSACCION_ORGANISMO','DURATION']

## Agregar los encabezados a la hoja de Excel
hoja2.append(column_headers)

## Iterar a través de las filas del DataFrame y agregar cada fila como una lista
for index, row in list_mayor_seven.iterrows():
    hoja2.append([row['ID_TRANSACCION_ORGANISMO'],row['DURATION']])

## hoja 3 Donde se guardaran los datos que se requieren para llenar el documento de graficas semanales
    
num_tarjetas = tarjetas_unicas.size                                                      
hoja3['A1'] = "informacion para Graficas Semanales"
hoja3.append(['## Tarjetas','## Transacciones','$ Promedio'])
hoja3.append([num_tarjetas,num_trx,mean_montos])

### Guardado de Archivos ##
## Guardar el DataFrame de resultados en un archivo CSV este contiene el Resumen del total de transacciones y montos por fecha
archivo_ex = f"Resumen_RRE_{quincena}.csv"
ruta_resultados = os.path.join(ruta_guardado, archivo_ex)
resultados.to_csv(ruta_resultados, index=False)

## Guardar el DataFrame completo en un archivo CSV este contiene la concatenacion de los archivos
## con extencion -Transacciones.csv el cual no tiene ninguna modificacion
ruta_completa = os.path.join(ruta_guardado, mes_completo)
df_completo.to_csv(ruta_completa, index=False)

## Este es el archivo en el cual se guardara el numero de transacciones mayores a 7 segundos
## las tablas de total de transacciones con su porcentaje ocupado por fisicas y digitales
## las tablas con el porcentaje a pagar segun el monto alcanzado
ruta_mp = os.path.join(ruta_guardado,archivo_mp)
wb.save(ruta_mp)

ruta_tr = os.path.join(ruta_guardado,f"Transacciones_{mes_nombre}.csv")
resumen.sort_index().to_csv(ruta_tr)

print("Proceso realizado con Exito!!")